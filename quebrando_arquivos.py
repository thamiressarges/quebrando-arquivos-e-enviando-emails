from openpyxl import load_workbook, Workbook

caminho_arquivo = r'./dados/Quebrar.xlsx'
planilha_aberta = load_workbook(caminho_arquivo)

sheet_selecionada = planilha_aberta['Dados']

arquivos_vendedores = {}

for linha in range(2, len(sheet_selecionada['A']) + 1):
    vendedor = sheet_selecionada[f'A{linha}'].value
    produto = sheet_selecionada[f'B{linha}'].value
    vendas = sheet_selecionada[f'C{linha}'].value

    if vendedor is None:
        continue 

    if vendedor not in arquivos_vendedores:
        
        novo_wb = Workbook()
        novo_ws = novo_wb.active
        novo_ws.title = "Resumo"

        novo_ws['A1'] = 'Vendedor'
        novo_ws['B1'] = 'Produto'
        novo_ws['C1'] = 'Vendas'

        arquivos_vendedores[vendedor] = novo_wb

    novo_wb = arquivos_vendedores[vendedor]
    novo_ws = novo_wb.active

    proxima_linha = novo_ws.max_row + 1

    novo_ws[f'A{proxima_linha}'] = vendedor
    novo_ws[f'B{proxima_linha}'] = produto
    novo_ws[f'C{proxima_linha}'] = vendas

for vendedor, wb in arquivos_vendedores.items():
    nome_arquivo = vendedor.replace(" ", "_").replace("/", "-")
    caminho_novo = fr'.\relatorios\{nome_arquivo}.xlsx'
    wb.save(caminho_novo)

